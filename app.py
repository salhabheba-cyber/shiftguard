"""ShiftGuard — Main Flask Application"""
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from flask_cors import CORS
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from functools import wraps
from datetime import datetime, timedelta, date
import logging, os, sys, shutil, io, base64, calendar, re
from logging.handlers import RotatingFileHandler
import config, database, network_manager

app = Flask(__name__)
app.config['SECRET_KEY'] = config.SECRET_KEY
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=config.SESSION_HOURS)
app.config['MAX_CONTENT_LENGTH'] = config.PHOTO_MAX_MB * 1024 * 1024
CORS(app)

# ── LOGGING ───────────────────────────────────────────────────────────────────
logger = logging.getLogger('shiftguard')
logger.setLevel(logging.INFO)
fh = RotatingFileHandler(config.LOG_FILE, maxBytes=10_485_760, backupCount=5)
fh.setFormatter(logging.Formatter('%(asctime)s %(levelname)s %(message)s'))
logger.addHandler(fh)
logger.addHandler(logging.StreamHandler(sys.stdout))

# ── AUTH ──────────────────────────────────────────────────────────────────────
login_manager = LoginManager(app)
login_manager.login_view = 'login_page'

class AdminUser(UserMixin):
    def __init__(self, d):
        self.id        = d['username']
        self.username  = d['username']
        self.role      = d.get('role','admin')
        self.branch_id = d.get('branch_id')

@login_manager.user_loader
def load_user(username):
    d = database.get_admin(username)
    return AdminUser(d) if d else None

def api_login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if not current_user.is_authenticated:
            return jsonify({'success':False,'message':'Login required'}), 401
        return f(*a, **kw)
    return dec

# ── SECURITY HEADERS ──────────────────────────────────────────────────────────
@app.after_request
def sec_headers(r):
    r.headers['X-Content-Type-Options'] = 'nosniff'
    r.headers['X-XSS-Protection']       = '1; mode=block'
    r.headers['X-Frame-Options']         = 'SAMEORIGIN'
    return r

# ── INIT ──────────────────────────────────────────────────────────────────────
with app.app_context():
    database.init_db()
    if not database.get_admin('admin'):
        database.create_admin('admin','admin123',role='superadmin')
        logger.warning("Default admin created: admin / admin123")

# ── PHOTO HELPER ──────────────────────────────────────────────────────────────
def save_photo(b64, emp_name, action):
    try:
        if ',' in b64: b64 = b64.split(',')[1]
        data   = base64.b64decode(b64)
        now    = datetime.now()
        safe   = re.sub(r'[^a-zA-Z0-9_]','_', emp_name)
        folder = os.path.join(config.PHOTOS_DIR, str(now.year), f'{now.month:02d}', f'{now.day:02d}')
        os.makedirs(folder, exist_ok=True)
        fname  = f"{safe}_{action}_{now.strftime('%H%M%S')}.jpg"
        fpath  = os.path.join(folder, fname)
        with open(fpath,'wb') as f: f.write(data)
        return os.path.relpath(fpath, os.path.join(config.BASE_DIR,'static')).replace('\\','/')
    except Exception as e:
        logger.error(f"Photo save error: {e}")
        return None

# ══════════════════════════════════════════════════════════════════════════════
# KIOSK
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/')
def home(): return redirect(url_for('kiosk'))

@app.route('/kiosk')
def kiosk():
    theme    = database.get_theme()
    branches = database.get_branches()
    logs     = database.get_today_attendance()
    checked  = len([l for l in logs if l['check_in'] and not l['check_out']])
    today    = date.today().strftime('%A, %d %B %Y')
    return render_template('kiosk.html', theme=theme, branches=branches,
                           checked_in_count=checked, today=today)

@app.route('/api/kiosk/search')
def kiosk_search():
    q         = request.args.get('q','').strip()
    branch_id = request.args.get('branch_id', type=int)
    if not q: return jsonify({'employees':[]})
    emps = database.search_employees(q, branch_id)
    return jsonify({'employees':[{'id':e['id'],'name':e['name'],
        'position':e['position'] or '','shift_start':e['shift_start'],
        'shift_end':e['shift_end']} for e in emps]})

@app.route('/api/kiosk/checkin', methods=['POST'])
def kiosk_checkin():
    d    = request.json or {}
    eid  = d.get('employee_id')
    pin  = str(d.get('pin','')).strip()
    if not eid or not pin:
        return jsonify({'success':False,'message':'Missing data'}), 400
    if not database.verify_employee_pin(eid, pin):
        return jsonify({'success':False,'message':'❌ Incorrect PIN. Try again.'}), 401
    emp   = database.get_employee(eid)
    photo = None
    pstat = d.get('photo_status','ok')
    if d.get('photo'):
        photo = save_photo(d['photo'], emp['name'], 'in')
        if not photo: pstat = 'save_failed'
    ok, msg = database.record_check_in(eid, photo, pstat)
    return jsonify({'success':ok,'message':msg,
                    'photo_path':f'/static/{photo}' if photo else None,
                    'time':datetime.now().strftime('%H:%M')})

@app.route('/api/kiosk/checkout', methods=['POST'])
def kiosk_checkout():
    d    = request.json or {}
    eid  = d.get('employee_id')
    pin  = str(d.get('pin','')).strip()
    if not eid or not pin:
        return jsonify({'success':False,'message':'Missing data'}), 400
    if not database.verify_employee_pin(eid, pin):
        return jsonify({'success':False,'message':'❌ Incorrect PIN. Try again.'}), 401
    emp   = database.get_employee(eid)
    photo = None
    pstat = d.get('photo_status','ok')
    if d.get('photo'):
        photo = save_photo(d['photo'], emp['name'], 'out')
        if not photo: pstat = 'save_failed'
    ok, msg = database.record_check_out(eid, photo, pstat)
    return jsonify({'success':ok,'message':msg,
                    'photo_path':f'/static/{photo}' if photo else None,
                    'time':datetime.now().strftime('%H:%M')})

@app.route('/api/kiosk/status')
def kiosk_status():
    bid  = request.args.get('branch_id', type=int)
    logs = database.get_today_attendance(bid)
    return jsonify({'checked_in_count': len([l for l in logs if l['check_in'] and not l['check_out']])})

# ══════════════════════════════════════════════════════════════════════════════
# ADMIN AUTH
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/admin/login', methods=['GET'])
def login_page():
    return render_template('login.html', theme=database.get_theme())

@app.route('/admin/login', methods=['POST'])
def login_post():
    d    = request.json or {}
    user = database.verify_admin(d.get('username',''), d.get('password',''))
    if not user:
        database.log_event('login_fail', d.get('username',''), request.remote_addr, 'Wrong password')
        return jsonify({'success':False,'message':'Invalid username or password'}), 401
    login_user(AdminUser(user), remember=True)
    database.log_event('login_ok', user['username'], request.remote_addr, '')
    return jsonify({'success':True})

@app.route('/admin/logout', methods=['POST'])
@login_required
def logout():
    database.log_event('logout', current_user.username, request.remote_addr, '')
    logout_user()
    return jsonify({'success':True})

# ══════════════════════════════════════════════════════════════════════════════
# ADMIN DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/admin')
@login_required
def admin():
    return render_template('admin.html', theme=database.get_theme(),
                           branches=database.get_branches(), user=current_user)

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
@app.route('/api/admin/dashboard')
@api_login_required
def api_dashboard():
    bid  = request.args.get('branch_id', type=int)
    logs = database.get_today_attendance(bid)
    emps = database.get_employees(active_only=True, branch_id=bid)
    checked_in  = [l for l in logs if l['check_in'] and not l['check_out']]
    checked_out = [l for l in logs if l['check_in'] and l['check_out']]
    late        = [l for l in logs if (l['minutes_late'] or 0) > config.LATE_THRESHOLD_MINUTES]
    absent      = max(0, len(emps) - len([l for l in logs if l['check_in']]))
    return jsonify({
        'total_employees': len(emps),
        'checked_in':  len(checked_in),
        'checked_out': len(checked_out),
        'late':        len(late),
        'absent':      absent,
        'today':       date.today().strftime('%A, %d %B %Y'),
        'logs': [{
            'id':              l['id'],
            'name':            l['name'],
            'position':        l['position'] or '',
            'branch_name':     l.get('branch_name',''),
            'check_in':        str(l['check_in'])[:16] if l['check_in'] else None,
            'check_in_photo':  f"/static/{l['check_in_photo']}" if l['check_in_photo'] else None,
            'check_out':       str(l['check_out'])[:16] if l['check_out'] else None,
            'check_out_photo': f"/static/{l['check_out_photo']}" if l['check_out_photo'] else None,
            'status':          l['status'],
            'minutes_late':    l['minutes_late'] or 0,
            'hours_worked':    round(l['hours_worked'] or 0, 2),
        } for l in logs]
    })

# ── ATTENDANCE ─────────────────────────────────────────────────────────────────
@app.route('/api/admin/attendance')
@api_login_required
def api_attendance():
    start = request.args.get('start', (date.today()-timedelta(days=30)).isoformat())
    end   = request.args.get('end',   date.today().isoformat())
    eid   = request.args.get('employee_id', type=int)
    bid   = request.args.get('branch_id',   type=int)
    logs  = database.get_attendance_range(start, end, eid, bid)
    return jsonify({'logs':[{
        'id':              l['id'],
        'name':            l['name'],
        'position':        l['position'] or '',
        'branch_name':     l.get('branch_name',''),
        'date':            l['date'],
        'check_in':        str(l['check_in'])[:19]  if l['check_in']  else '',
        'check_in_photo':  f"/static/{l['check_in_photo']}"  if l['check_in_photo']  else None,
        'check_out':       str(l['check_out'])[:19] if l['check_out'] else '',
        'check_out_photo': f"/static/{l['check_out_photo']}" if l['check_out_photo'] else None,
        'status':          l['status'] or '',
        'minutes_late':    l['minutes_late'] or 0,
        'hours_worked':    round(l['hours_worked'] or 0, 2),
        'overtime_hours':  round(l['overtime_hours'] or 0, 2),
        'notes':           l['notes'] or '',
    } for l in logs], 'count':len(logs)})

@app.route('/api/admin/attendance/<int:rid>', methods=['PUT'])
@api_login_required
def api_edit_att(rid):
    d = request.json or {}
    database.admin_edit_attendance(rid, d.get('check_in') or None,
                                   d.get('check_out') or None, d.get('notes'),
                                   current_user.username)
    return jsonify({'success':True,'message':'Record updated'})

@app.route('/api/admin/attendance/<int:rid>', methods=['DELETE'])
@api_login_required
def api_del_att(rid):
    database.admin_delete_attendance(rid)
    return jsonify({'success':True})

@app.route('/api/admin/attendance/add', methods=['POST'])
@api_login_required
def api_add_att():
    d  = request.json or {}
    ok = database.admin_add_attendance(d.get('employee_id'), d.get('date'),
                                       d.get('check_in'), d.get('check_out'), d.get('notes'))
    return jsonify({'success':ok})

@app.route('/api/admin/attendance/<int:rid>/delete-photo', methods=['POST'])
@api_login_required
def api_del_photo(rid):
    d     = request.json or {}
    field = d.get('field','check_in_photo')
    r     = database.q('SELECT * FROM attendance WHERE id=?',(rid,))
    if r and r[0].get(field):
        path = os.path.join(config.BASE_DIR,'static',r[0][field])
        if os.path.exists(path): os.remove(path)
    database.delete_photo(rid, field)
    return jsonify({'success':True})

# ── EMPLOYEES ──────────────────────────────────────────────────────────────────
@app.route('/api/admin/employees', methods=['GET'])
@api_login_required
def api_employees():
    bid  = request.args.get('branch_id', type=int)
    emps = database.get_employees(active_only=True, branch_id=bid)
    for e in emps: e.pop('pin_hash', None)
    return jsonify({'employees': emps})

@app.route('/api/admin/employees', methods=['POST'])
@api_login_required
def api_add_emp():
    d    = request.json or {}
    name = (d.get('name') or '').strip()
    pin  = str(d.get('pin','')).strip()
    if not name: return jsonify({'success':False,'message':'Name required'}), 400
    if not pin or len(pin)!=4 or not pin.isdigit():
        return jsonify({'success':False,'message':'PIN must be exactly 4 digits'}), 400
    try:
        eid = database.add_employee(
            name=name, pin=pin,
            position=d.get('position') or None, phone=d.get('phone') or None,
            branch_id=int(d.get('branch_id') or 1),
            shift_start=d.get('shift_start','09:30'), shift_end=d.get('shift_end','18:30'),
            working_days=d.get('working_days','Mon,Tue,Wed,Thu,Fri,Sat'),
            salary_type=d.get('salary_type','monthly'),
            salary_amount=float(d.get('salary_amount') or 0),
            hire_date=d.get('hire_date') or None)
        return jsonify({'success':True,'message':f'{name} added!','id':eid})
    except Exception as e:
        return jsonify({'success':False,'message':str(e)}), 500

@app.route('/api/admin/employees/<int:eid>', methods=['PUT'])
@api_login_required
def api_upd_emp(eid):
    d  = request.json or {}
    kw = {}
    for f in ['name','position','phone','branch_id','shift_start','shift_end',
              'working_days','salary_type','salary_amount','hire_date','is_active']:
        if f in d: kw[f] = d[f]
    if 'pin' in d and d['pin']:
        pin = str(d['pin']).strip()
        if len(pin)!=4 or not pin.isdigit():
            return jsonify({'success':False,'message':'PIN must be 4 digits'}), 400
        kw['pin'] = pin
    database.update_employee(eid, **kw)
    return jsonify({'success':True,'message':'Employee updated'})

@app.route('/api/admin/employees/<int:eid>/reset-pin', methods=['POST'])
@api_login_required
def api_reset_pin(eid):
    d   = request.json or {}
    pin = str(d.get('pin','')).strip()
    if len(pin)!=4 or not pin.isdigit():
        return jsonify({'success':False,'message':'PIN must be 4 digits'}), 400
    database.reset_pin(eid, pin)
    emp = database.get_employee(eid)
    return jsonify({'success':True,'message':f"PIN reset for {emp['name'] if emp else 'employee'}"})

@app.route('/api/admin/employees/<int:eid>', methods=['DELETE'])
@api_login_required
def api_del_emp(eid):
    database.delete_employee(eid)
    return jsonify({'success':True,'message':'Employee permanently deleted'})

# ── BRANCHES ───────────────────────────────────────────────────────────────────
@app.route('/api/admin/branches', methods=['GET'])
@api_login_required
def api_branches():
    return jsonify({'branches': database.get_branches(active_only=False)})

@app.route('/api/admin/branches', methods=['POST'])
@api_login_required
def api_add_branch():
    d    = request.json or {}
    name = (d.get('name') or '').strip()
    if not name: return jsonify({'success':False,'message':'Name required'}), 400
    bid  = database.add_branch(name, d.get('address'), d.get('phone'))
    return jsonify({'success':True,'id':bid,'message':f'Branch "{name}" added'})

@app.route('/api/admin/branches/<int:bid>', methods=['PUT'])
@api_login_required
def api_upd_branch(bid):
    d = request.json or {}
    database.update_branch(bid, d.get('name',''), d.get('address'), d.get('phone'))
    return jsonify({'success':True})

@app.route('/api/admin/branches/<int:bid>', methods=['DELETE'])
@api_login_required
def api_del_branch(bid):
    database.delete_branch(bid)
    return jsonify({'success':True})

# ── SALARY ─────────────────────────────────────────────────────────────────────
@app.route('/api/admin/salary/calculate', methods=['POST'])
@api_login_required
def api_calc():
    d     = request.json or {}
    year  = int(d.get('year',  date.today().year))
    month = int(d.get('month', date.today().month))
    eid   = d.get('employee_id')
    emps  = [database.get_employee(int(eid))] if eid else database.get_employees()
    recs  = []
    for e in emps:
        if not e: continue
        calc = database.calculate_salary(e['id'], year, month)
        database.save_salary(calc)
        recs.append(calc)
    return jsonify({'success':True,'records':recs})

@app.route('/api/admin/salary/records')
@api_login_required
def api_sal_records():
    return jsonify({'records': database.get_salary_records(
        request.args.get('year',type=int),
        request.args.get('month',type=int),
        request.args.get('employee_id',type=int))})

@app.route('/api/admin/salary/mark-paid', methods=['POST'])
@api_login_required
def api_mark_paid():
    d = request.json or {}
    database.mark_paid(d['employee_id'], d['year'], d['month'])
    return jsonify({'success':True})

# ── BLOCKED SITES ──────────────────────────────────────────────────────────────
@app.route('/api/admin/blocked-sites', methods=['GET'])
@api_login_required
def api_blocked():
    return jsonify({'sites': database.get_blocked_sites()})

@app.route('/api/admin/blocked-sites', methods=['POST'])
@api_login_required
def api_add_blocked():
    d  = request.json or {}
    ok = database.add_blocked_site(d.get('domain',''), d.get('category','restricted'))
    return jsonify({'success':ok,'message':'Added' if ok else 'Already exists'})

@app.route('/api/admin/blocked-sites/<int:sid>', methods=['DELETE'])
@api_login_required
def api_del_blocked(sid):
    database.remove_blocked_site(sid)
    return jsonify({'success':True})

# ── NETWORK SECURITY ───────────────────────────────────────────────────────────
@app.route('/api/admin/network/info')
@api_login_required
def api_net_info():
    return jsonify({'info':network_manager.get_network_info(),
                    'hosts_status':network_manager.get_hosts_status()})

@app.route('/api/admin/network/scan')
@api_login_required
def api_net_scan():
    devices = network_manager.scan_network()
    return jsonify({'devices':devices,'count':len(devices)})

@app.route('/api/admin/network/apply-hosts', methods=['POST'])
@api_login_required
def api_apply_hosts():
    sites   = database.get_blocked_sites()
    domains = [s['domain'] for s in sites if s['is_active']]
    ok, msg = network_manager.apply_hosts_blocking(domains)
    if ok: database.log_event('hosts_applied', current_user.username, request.remote_addr, f'{len(domains)} domains')
    return jsonify({'success':ok,'message':msg})

@app.route('/api/admin/network/remove-hosts', methods=['POST'])
@api_login_required
def api_remove_hosts():
    ok, msg = network_manager.remove_hosts_blocking()
    if ok: database.log_event('hosts_removed', current_user.username, request.remote_addr, '')
    return jsonify({'success':ok,'message':msg})

# ── THEME / SETTINGS ───────────────────────────────────────────────────────────
@app.route('/api/admin/theme', methods=['GET'])
@api_login_required
def api_get_theme():
    return jsonify({'theme':database.get_theme()})

@app.route('/api/admin/theme', methods=['POST'])
@api_login_required
def api_save_theme():
    d = request.json or {}
    for k in ['primary','secondary','accent','dark','text','font','logo_text','dark_mode']:
        if k in d: database.set_setting(f'theme_{k}', str(d[k]).replace('#',''))
    return jsonify({'success':True,'message':'Theme saved'})

@app.route('/api/admin/change-password', methods=['POST'])
@api_login_required
def api_change_pw():
    d    = request.json or {}
    ok, msg = database.change_password(current_user.username,
                                       d.get('current_password',''), d.get('new_password',''))
    if ok: logout_user()
    return jsonify({'success':ok,'message':msg})

# ── ADMIN USERS ────────────────────────────────────────────────────────────────
@app.route('/api/admin/users', methods=['GET'])
@api_login_required
def api_users():
    return jsonify({'users':database.get_all_admins()})

@app.route('/api/admin/users', methods=['POST'])
@api_login_required
def api_add_user():
    d = request.json or {}
    try:
        database.create_admin(d['username'],d['password'],d.get('role','admin'),d.get('branch_id'))
        return jsonify({'success':True,'message':'User created'})
    except Exception as e:
        return jsonify({'success':False,'message':str(e)}), 500

# ── LOGS ───────────────────────────────────────────────────────────────────────
@app.route('/api/admin/logs')
@api_login_required
def api_logs():
    return jsonify({'logs':database.get_logs()})

# ── PHOTOS ─────────────────────────────────────────────────────────────────────
@app.route('/api/admin/photos')
@api_login_required
def api_photos():
    start = request.args.get('start',(date.today()-timedelta(days=30)).isoformat())
    end   = request.args.get('end',  date.today().isoformat())
    logs  = database.get_attendance_range(start, end)
    photos= []
    for l in logs:
        for field,label in [('check_in_photo','Check In'),('check_out_photo','Check Out')]:
            if l.get(field):
                photos.append({'att_id':l['id'],'field':field,'employee':l['name'],
                                'date':l['date'],'type':label,'url':f"/static/{l[field]}"})
    return jsonify({'photos':photos})

# ── BACKUP ─────────────────────────────────────────────────────────────────────
@app.route('/api/admin/backup')
@api_login_required
def api_backup():
    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    dest = os.path.join(config.BACKUP_DIR, f'shiftguard_backup_{ts}.db')
    shutil.copy2(config.DATABASE_PATH, dest)
    return send_file(dest, as_attachment=True, download_name=f'shiftguard_backup_{ts}.db')

# ── REPORTS ────────────────────────────────────────────────────────────────────
@app.route('/api/admin/report/daily')
@api_login_required
def api_daily_report():
    day = request.args.get('date', date.today().isoformat())
    bid = request.args.get('branch_id', type=int)
    return _make_report('daily', day=day, branch_id=bid)

@app.route('/api/admin/report/monthly')
@api_login_required
def api_monthly_report():
    month_str = request.args.get('month', date.today().strftime('%Y-%m'))
    bid       = request.args.get('branch_id', type=int)
    return _make_report('monthly', month_str=month_str, branch_id=bid)

def _make_report(rtype, **kw):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        T,DK,TL,WH='00A8A8','1c3a3b','E8F8F8','FFFFFF'
        def H(cell, val, bg=None):
            cell.value=val; cell.font=Font(bold=True,color=WH,size=10)
            cell.fill=PatternFill('solid',fgColor=bg or T)
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        wb=openpyxl.Workbook()
        if rtype=='daily':
            day=kw['day']; bid=kw.get('branch_id')
            logs=database.get_attendance_range(day,day,branch_id=bid)
            ws=wb.active; ws.title=f'Daily {day}'
            ws.merge_cells('A1:I1'); H(ws['A1'],f'ShiftGuard — Daily Report — {day}',DK)
            ws.row_dimensions[1].height=28
            for c,h in enumerate(['Employee','Branch','Position','Date','Check In',
                                   'Check Out','Hours','Late(min)','Status'],1):
                H(ws.cell(row=2,column=c),h)
            for i,l in enumerate(logs,3):
                row=[l['name'],l.get('branch_name',''),l['position'] or '',l['date'],
                     str(l['check_in'])[:16] if l['check_in'] else '—',
                     str(l['check_out'])[:16] if l['check_out'] else '—',
                     round(l['hours_worked'] or 0,2),l['minutes_late'] or 0,l['status'] or '—']
                for c,v in enumerate(row,1):
                    cell=ws.cell(row=i,column=c); cell.value=v
                    cell.alignment=Alignment(horizontal='center')
                    if i%2==0: cell.fill=PatternFill('solid',fgColor=TL)
            for i,w in enumerate([22,16,14,12,18,18,10,10,14],1):
                ws.column_dimensions[get_column_letter(i)].width=w
            fname=f'ShiftGuard_Daily_{day}.xlsx'
        else:
            ms=kw['month_str']; bid=kw.get('branch_id')
            year,month=map(int,ms.split('-'))
            mname=calendar.month_name[month]
            first=f'{year}-{month:02d}-01'
            last =f'{year}-{month:02d}-{calendar.monthrange(year,month)[1]}'
            emps=database.get_employees(active_only=True)
            logs=database.get_attendance_range(first,last,branch_id=bid)
            # Sheet 1 Summary
            ws1=wb.active; ws1.title='Attendance Summary'
            ws1.merge_cells('A1:K1'); H(ws1['A1'],f'ShiftGuard — Monthly Attendance — {mname} {year}',DK)
            ws1.row_dimensions[1].height=28
            for c,h in enumerate(['Employee','Branch','Scheduled','Present','Absent','Late Days',
                                   'Total Hours','Overtime','Late Min','Status','Notes'],1):
                H(ws1.cell(row=2,column=c),h)
            for i,e in enumerate(emps,3):
                calc=database.calculate_salary(e['id'],year,month)
                row=[e['name'],e.get('branch_name',''),calc['working_days'],calc['days_present'],
                     calc['days_absent'],calc['days_late'],calc['total_hours'],
                     calc['overtime_hours'],calc['total_late_min'],'Active','']
                for c,v in enumerate(row,1):
                    cell=ws1.cell(row=i,column=c); cell.value=v
                    cell.alignment=Alignment(horizontal='center')
                    if i%2==0: cell.fill=PatternFill('solid',fgColor=TL)
            for i,w in enumerate([22,16,12,10,10,12,12,12,10,10,14],1):
                ws1.column_dimensions[get_column_letter(i)].width=w
            # Sheet 2 Salary
            ws2=wb.create_sheet('Salary')
            ws2.merge_cells('A1:H1'); H(ws2['A1'],f'Salary Breakdown — {mname} {year}',DK)
            for c,h in enumerate(['Employee','Base','Absent Ded.','Late Ded.','OT Pay','Net Salary','Paid','Date'],1):
                H(ws2.cell(row=2,column=c),h)
            for i,e in enumerate(emps,3):
                calc=database.calculate_salary(e['id'],year,month)
                row=[e['name'],calc['base_salary'],calc['absent_deduction'],
                     calc['late_deduction'],calc['overtime_pay'],calc['net_salary'],'','']
                for c,v in enumerate(row,1):
                    cell=ws2.cell(row=i,column=c); cell.value=v
                    cell.alignment=Alignment(horizontal='center')
                    if c==6: cell.font=Font(bold=True)
                    if i%2==0: cell.fill=PatternFill('solid',fgColor=TL)
            for i,w in enumerate([22,14,14,12,14,14,8,14],1):
                ws2.column_dimensions[get_column_letter(i)].width=w
            # Sheet 3 Daily Log
            ws3=wb.create_sheet('Daily Log')
            ws3.merge_cells('A1:H1'); H(ws3['A1'],f'Daily Log — {mname} {year}',DK)
            for c,h in enumerate(['Employee','Date','Check In','Check Out','Hours','OT','Late(min)','Status'],1):
                H(ws3.cell(row=2,column=c),h)
            for i,l in enumerate(logs,3):
                row=[l['name'],l['date'],
                     str(l['check_in'])[:16] if l['check_in'] else '—',
                     str(l['check_out'])[:16] if l['check_out'] else '—',
                     round(l['hours_worked'] or 0,2),round(l['overtime_hours'] or 0,2),
                     l['minutes_late'] or 0,l['status'] or '—']
                for c,v in enumerate(row,1):
                    cell=ws3.cell(row=i,column=c); cell.value=v
                    cell.alignment=Alignment(horizontal='center')
                    if i%2==0: cell.fill=PatternFill('solid',fgColor=TL)
            fname=f'ShiftGuard_Monthly_{ms}.xlsx'
        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,download_name=fname)
    except Exception as e:
        logger.error(f"Report error: {e}")
        return jsonify({'error':str(e)}),500

# ── ERRORS ─────────────────────────────────────────────────────────────────────
@app.errorhandler(404)
def e404(e): return render_template('error.html',error='Page not found',theme=database.get_theme()),404
@app.errorhandler(500)
def e500(e): return render_template('error.html',error='Server error',theme=database.get_theme()),500

# ── START ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    logger.info(f"Starting ShiftGuard on port {config.PORT}")
    try:
        from waitress import serve
        logger.info("Production server: Waitress")
        serve(app, host=config.HOST, port=config.PORT, threads=4)
    except ImportError:
        logger.warning("Waitress not found, using Flask dev server")
        app.run(host=config.HOST, port=config.PORT, debug=False, threaded=True)
